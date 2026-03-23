"""Helpers to download Pipefy automation job exports and convert the first sheet to CSV."""

from __future__ import annotations

import csv
import io
from typing import Final
from urllib.parse import urlparse

import httpx
from openpyxl import load_workbook

_ALLOWED_HOST_SUFFIX: Final[str] = ".pipefy.com"

_DEFAULT_DOWNLOAD_TIMEOUT: Final[int] = 120


def is_allowed_pipefy_export_download_url(url: str) -> bool:
    """Return True if the URL is an https URL on a Pipefy host (signed export links).

    Args:
        url: Absolute URL from ``automationJobsExport.fileUrl``.
    """
    parsed = urlparse(url.strip())
    if parsed.scheme != "https":
        return False
    host = (parsed.hostname or "").lower()
    return host.endswith(_ALLOWED_HOST_SUFFIX)


def xlsx_first_sheet_to_csv_limited(
    xlsx_bytes: bytes,
    *,
    max_output_chars: int,
) -> tuple[str, int, str, bool]:
    """Convert the first worksheet of an xlsx workbook to CSV text with a character cap.

    Args:
        xlsx_bytes: Raw .xlsx file contents.
        max_output_chars: Maximum UTF-8 character length of the returned CSV (excluding
            a possible truncation notice line).

    Returns:
        Tuple of (csv_text, rows_included, sheet_title, truncated).
    """
    if max_output_chars < 1:
        raise ValueError("max_output_chars must be at least 1.")

    buf = io.BytesIO(xlsx_bytes)
    wb = load_workbook(buf, read_only=True, data_only=True)
    try:
        ws = wb.active
        title = str(ws.title) if ws.title else ""
        out = io.StringIO()
        total_chars = 0
        rows_included = 0
        truncated = False

        for row in ws.iter_rows(values_only=True):
            line_buf = io.StringIO()
            line_writer = csv.writer(line_buf)
            line_writer.writerow(["" if c is None else c for c in row])
            piece = line_buf.getvalue()
            if total_chars + len(piece) > max_output_chars:
                remaining = max_output_chars - total_chars
                if remaining > 0:
                    out.write(piece[:remaining])
                truncated = True
                break
            out.write(piece)
            total_chars += len(piece)
            rows_included += 1

        text = out.getvalue()
        if truncated:
            text += "\n# [truncated by pipefy-mcp-server: max_output_chars exceeded]\n"
        return text, rows_included, title, truncated
    finally:
        wb.close()


async def download_bytes(url: str, *, max_bytes: int) -> bytes:
    """GET a URL and return the body, enforcing a size limit.

    Args:
        url: HTTPS URL to fetch.
        max_bytes: Abort with ``ValueError`` if Content-Length or streamed size exceeds this.

    Raises:
        ValueError: On disallowed URL, oversize body, or non-success status.
        httpx.HTTPError: On transport errors.
    """
    if not is_allowed_pipefy_export_download_url(url):
        raise ValueError("Download URL is not an allowed Pipefy https URL.")

    timeout = httpx.Timeout(_DEFAULT_DOWNLOAD_TIMEOUT)
    limits = httpx.Limits(max_keepalive_connections=5, max_connections=5)
    async with httpx.AsyncClient(
        timeout=timeout, limits=limits, follow_redirects=True
    ) as client:
        async with client.stream("GET", url) as response:
            response.raise_for_status()
            cl = response.headers.get("content-length")
            if cl is not None:
                try:
                    declared = int(cl)
                except ValueError:
                    pass
                else:
                    if declared > max_bytes:
                        raise ValueError(
                            f"Export file exceeds max_download_bytes ({max_bytes} bytes)."
                        )
            chunks: list[bytes] = []
            total = 0
            async for chunk in response.aiter_bytes():
                total += len(chunk)
                if total > max_bytes:
                    raise ValueError(
                        f"Export file exceeds max_download_bytes ({max_bytes} bytes)."
                    )
                chunks.append(chunk)
            return b"".join(chunks)


__all__ = [
    "download_bytes",
    "is_allowed_pipefy_export_download_url",
    "xlsx_first_sheet_to_csv_limited",
]
